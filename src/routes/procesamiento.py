from flask import Blueprint, request, jsonify, send_file
import pandas as pd
import requests
import time
from sklearn.cluster import KMeans
import math
import os
import tempfile
import matplotlib.pyplot as plt
import io
import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
import traceback

procesamiento_bp = Blueprint("procesamiento", __name__)

GEOAPIFY_API_KEY = "65de779bc48c42d8a1208a5f5e9320b4"

# Coordenadas de la base (oficina/depósito) para el cálculo de distancias
BASE_LAT = -34.6  # Ejemplo: Latitud de Buenos Aires
BASE_LNG = -58.4  # Ejemplo: Longitud de Buenos Aires

# --- FUNCIÓN DE ESTILOS MEJORADA ---
def apply_styles_and_autofit(writer, sheet_name, df):
    """
    Aplica formato (moneda, decimales, negrita, centrado) y autoajusta el ancho de las columnas
    a una hoja de Excel específica.
    """
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    # --- Estilos ---
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    if 'currency' not in workbook.style_names:
        currency_style = NamedStyle(name='currency', number_format='$ #,##0.00', alignment=center_alignment)
        workbook.add_named_style(currency_style)
    if 'decimal' not in workbook.style_names:
        decimal_style = NamedStyle(name='decimal', number_format='0.00', alignment=center_alignment)
        workbook.add_named_style(decimal_style)
    if 'default_center' not in workbook.style_names:
        default_center_style = NamedStyle(name='default_center', alignment=center_alignment)
        workbook.add_named_style(default_center_style)

    bold_font = Font(bold=True)
    header_font = Font(bold=True)
    
    # Aplicar estilo al encabezado
    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = center_alignment
    
    # --- Mapeo de columnas a estilos ---
    style_map = {
        'Venta Total Anual': 'currency',
        'Score Oportunidad': 'decimal',
        'Score Prioridad': 'decimal',
        'Distancia Base (km)': 'decimal',
        'Latitud': 'decimal',
        'Longitud': 'decimal'
    }
    month_cols = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic", "Importe PS", "IMPORTE"]

    # Aplicar estilos a las celdas de datos
    for r_idx in range(2, worksheet.max_row + 1):
        for c_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=r_idx, column=c_idx)
            column_name = df.columns[c_idx - 1]
            
            # Aplicar estilo por defecto (centrado)
            cell.style = 'default_center'

            # Poner en negrita los nombres de cliente
            if column_name in ['Cliente', 'Razón Social', 'nombre']:
                cell.font = bold_font

            # Aplicar estilos de número específicos
            if column_name in style_map:
                cell.style = style_map[column_name]
            elif column_name in month_cols:
                cell.style = 'currency'

    # Autoajustar el ancho de las columnas
    for column_cells in worksheet.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width


# --- NUEVA FUNCIÓN DE ESTILO EXCLUSIVA PARA EL DASHBOARD ---
def style_dashboard_sheet(writer, sheet_name):
    """Aplica formato específico a la hoja de Dashboard."""
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    titles_to_bold = [
        "ESTADÍSTICAS GENERALES", "--- Distribución de Clientes ---",
        "--- Promedios ---", "TOP 5 CLIENTES POR PRIORIZACIÓN"
    ]
    
    # Poner títulos en negrita y centrar todo
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value in titles_to_bold:
                cell.font = bold_font
            cell.alignment = center_alignment # Centrar todas las celdas del dashboard

    # Autoajustar columnas del dashboard
    for column_cells in worksheet.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column].width = max_length + 2


def geocode_address_geoapify(address):
    url = f"https://api.geoapify.com/v1/geocode/search?text={address}&apiKey={GEOAPIFY_API_KEY}"
    try:
        response = requests.get(url)
        response.raise_for_status()
        data = response.json()
        if data and data["features"]:
            lon = data["features"][0]["geometry"]["coordinates"][0]
            lat = data["features"][0]["geometry"]["coordinates"][1]
            return lat, lon
        else:
            return None, None
    except requests.exceptions.RequestException as e:
        print(f"Error en la solicitud a Geoapify para {address}: {e}")
        return None, None
    except Exception as e:
        print(f"Error procesando la respuesta de Geoapify para {address}: {e}")
        return None, None

def geocode_addresses(df):
    if "Provincia" in df.columns:
        df["Direccion_Completa"] = df["Domicilio"] + ", " + df["Localidad"] + ", " + df["Provincia"]
    else:
        df["Direccion_Completa"] = df["Domicilio"] + ", " + df["Localidad"]
    
    latitudes = []
    longitudes = []
    
    for index, row in df.iterrows():
        address = row["Direccion_Completa"]
        lat, lon = geocode_address_geoapify(address)
        latitudes.append(lat)
        longitudes.append(lon)
        time.sleep(0.1)
    
    df["Latitud"] = latitudes
    df["Longitud"] = longitudes
    
    df_geocoded = df.dropna(subset=["Latitud", "Longitud"])
    
    return df_geocoded

def unify_data(df_clientes, df_ventas):
    month_columns = [col for col in df_ventas.columns if col in ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]]
    if "Importe PS" in df_ventas.columns:
        month_columns.append("Importe PS")
    elif "IMPORTE" in df_ventas.columns:
        month_columns.append("IMPORTE")

    for col in month_columns:
        df_ventas[col] = pd.to_numeric(df_ventas[col], errors="coerce").fillna(0)

    df_ventas_grouped = df_ventas.groupby("Cliente")[month_columns].sum().reset_index()

    if "Importe PS" in df_ventas.columns:
        df_ventas_grouped["Venta_Total_Anual"] = df_ventas_grouped["Importe PS"]
    elif "IMPORTE" in df_ventas.columns:
        df_ventas_grouped["Venta_Total_Anual"] = df_ventas_grouped["IMPORTE"]
    else:
        df_ventas_grouped["Venta_Total_Anual"] = df_ventas_grouped[month_columns].sum(axis=1)

    df_unificado = pd.merge(df_clientes, df_ventas_grouped, left_on="Razón Social", right_on="Cliente", how="left")
    df_unificado["Venta_Total_Anual"] = df_unificado["Venta_Total_Anual"].fillna(0)
    
    return df_unificado

def clasificar_abc(df):
    df_sorted = df.sort_values("Venta_Total_Anual", ascending=False).reset_index(drop=True)
    
    total_ventas = df_sorted["Venta_Total_Anual"].sum()
    if total_ventas == 0:
        df_sorted["Categoria_ABC"] = "C"
        return df_sorted
    
    df_sorted["venta_acumulada"] = df_sorted["Venta_Total_Anual"].cumsum()
    df_sorted["porcentaje_acumulado"] = df_sorted["venta_acumulada"] / total_ventas
    
    df_sorted["Categoria_ABC"] = df_sorted["porcentaje_acumulado"].apply(
        lambda x: "A" if x <= 0.8 else ("B" if x <= 0.95 else "C")
    )
    
    return df_sorted

def calcular_score_oportunidad(row):
    score = 0
    categoria_abc = row.get("Categoria_ABC", "C")
    if categoria_abc == "A": score += 40
    elif categoria_abc == "B": score += 25
    else: score += 10

    segmento = row.get("Segmento", "Otros/Sin datos")
    if segmento == "Distribuidor A": score += 30
    elif segmento == "Distribuidor B": score += 25
    elif segmento == "Mostrador A": score += 20
    elif segmento == "Mostrador B": score += 15
    else: score += 10

    rubro = row.get("Rubro", "Otros")
    if rubro == "Industrial": score += 20
    elif rubro == "Eléctrico": score += 18
    elif rubro == "Ferretero": score += 15
    elif rubro == "Repuestero": score += 12
    else: score += 10

    venta_total = row.get("Venta_Total_Anual", 0)
    if venta_total >= 10000000: score += 10
    elif venta_total >= 5000000: score += 8
    elif venta_total >= 1000000: score += 6
    else: score += 3

    return score

def calcular_frecuencia_visita(row):
    categoria = row.get("Categoria_ABC", "C")
    score_oportunidad = row.get("Score_Oportunidad", 0)

    if categoria == "A":
        return 7 if score_oportunidad >= 80 else 14
    elif categoria == "B":
        return 21
    else:
        return 30

def calcular_distancia(lat1, lng1, lat2, lng2):
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlng = math.radians(lng2 - lng1)
    a = (math.sin(dlat/2)**2 +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(dlng/2)**2)
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c

def calcular_prioridad_final(row):
    score_oportunidad_norm = row.get("Score_Oportunidad", 0) / 100
    
    if pd.isna(row.get("Latitud")) or pd.isna(row.get("Longitud")):
        distancia_desde_base = 999999
    else:
        distancia_desde_base = calcular_distancia(BASE_LAT, BASE_LNG, row["Latitud"], row["Longitud"])

    distancia_norm = 1.0 if distancia_desde_base == 0 else 1 / (1 + distancia_desde_base / 10)
    frecuencia_visita_dias = row.get("Frecuencia_Visita_Dias", 30)
    frecuencia_norm = 1.0 if frecuencia_visita_dias == 0 else 1 / (frecuencia_visita_dias / 30)

    peso_oportunidad = 0.5
    peso_distancia = 0.2
    peso_frecuencia = 0.3

    score_final = (score_oportunidad_norm * peso_oportunidad +
                   distancia_norm * peso_distancia +
                   frecuencia_norm * peso_frecuencia) * 100
    
    return score_final

def cluster_clients(df, num_clusters):
    df_valid = df.dropna(subset=["Latitud", "Longitud"])
    
    if len(df_valid) < num_clusters:
        num_clusters = max(1, len(df_valid))
    
    if len(df_valid) > 0:
        kmeans = KMeans(n_clusters=num_clusters, random_state=42, n_init=10)
        df_valid["Cluster"] = kmeans.fit_predict(df_valid[["Latitud", "Longitud"]])
        df = df.merge(df_valid[["Razón Social", "Cluster"]], on="Razón Social", how="left")
        df["Cluster"] = df["Cluster"].fillna(-1)
    else:
        df["Cluster"] = -1
    
    return df

def create_my_maps_sheet(df, writer):
    df_geocoded = df.dropna(subset=["Latitud", "Longitud"]).copy()
    
    if df_geocoded.empty:
        my_maps_data = pd.DataFrame(columns=["nombre", "latitud", "longitud", "cluster"])
    else:
        my_maps_data = pd.DataFrame({
            "nombre": df_geocoded["Razón Social"],
            "latitud": df_geocoded["Latitud"],
            "longitud": df_geocoded["Longitud"],
            "cluster": df_geocoded["Cluster"].apply(lambda x: f"Zona {int(x)+1}" if x != -1 else "Sin Zona")
        })
    
    my_maps_data.to_excel(writer, sheet_name="My Maps", index=False)
    apply_styles_and_autofit(writer, "My Maps", my_maps_data)
    return len(my_maps_data)

def create_dashboard_sheet(df, writer):
    total_clientes = len(df)
    abc_counts = df["Categoria_ABC"].value_counts()
    
    score_oportunidad_promedio = df["Score_Oportunidad"].mean() if "Score_Oportunidad" in df.columns else 0
    score_prioridad_promedio = df["Score_Prioridad_Final"].mean() if "Score_Prioridad_Final" in df.columns else 0
    
    df_geocoded_only = df.dropna(subset=["Latitud", "Longitud"])
    distancia_promedio = df_geocoded_only["Distancia_Desde_Base"].mean() if "Distancia_Desde_Base" in df_geocoded_only.columns else 0

    dashboard_data = [
        ["ESTADÍSTICAS GENERALES", ""], [""], ["Métrica", "Valor"],
        ["--- Distribución de Clientes ---", ""]
    ]
    
    for categoria in ["A", "B", "C"]:
        count = abc_counts.get(categoria, 0)
        percentage = (count / total_clientes * 100) if total_clientes > 0 else 0
        dashboard_data.append([f'Cantidad de Clientes "{categoria}"', f"{count} ({percentage:.1f}%)"])
    
    dashboard_data.extend([
        ["Total", f"{total_clientes}"], ["--- Promedios ---", ""],
        ["Score de Oportunidad Promedio", f"{score_oportunidad_promedio:.1f}"],
        ["Score de Prioridad Promedio", f"{score_prioridad_promedio:.1f}"],
        ["Distancia Promedio (km)", f"{distancia_promedio:.1f}"], [""],
        ["TOP 5 CLIENTES POR PRIORIZACIÓN", ""], ["Posición: Razón Social", "Score Prioridad Final"]
    ])

    if "Score_Prioridad_Final" in df.columns:
        top_5_clientes = df.sort_values(by="Score_Prioridad_Final", ascending=False).head(5)
        for i, (index, row) in enumerate(top_5_clientes.iterrows()):
            dashboard_data.append([f"{i+1}: {row['Razón Social']}", f"{row['Score_Prioridad_Final']:.2f}"])
    
    dashboard_df = pd.DataFrame(dashboard_data, columns=["Métrica", "Valor"])
    dashboard_df.to_excel(writer, sheet_name="Dashboard", index=False)
    
    # Llamar a la nueva función de estilo para el dashboard
    style_dashboard_sheet(writer, "Dashboard")


@procesamiento_bp.route("/procesar", methods=["POST"])
def procesar_datos():
    try:
        if "archivo_clientes" not in request.files or "archivo_ventas" not in request.files:
            return jsonify({"error": "Faltan archivos"}), 400
        
        archivo_clientes = request.files["archivo_clientes"]
        archivo_ventas = request.files["archivo_ventas"]
        num_clusters = int(request.form.get("num_clusters", 5))
        
        if archivo_clientes.filename == "" or archivo_ventas.filename == "":
            return jsonify({"error": "No se seleccionaron archivos"}), 400
        
        temp_dir = tempfile.mkdtemp()
        
        df_clientes = pd.read_excel(archivo_clientes)
        df_ventas = pd.read_excel(archivo_ventas, sheet_name="MIX POR CLIENTE")
        
        if "Unnamed: 0" in df_ventas.columns:
            df_ventas = df_ventas.rename(columns={"Unnamed: 0": "Cliente"})
        
        df_unificado = unify_data(df_clientes, df_ventas)
        df_clasificado = clasificar_abc(df_unificado)
        
        if 'Segmento' not in df_clasificado.columns:
            df_clasificado["Segmento"] = "Otros/Sin datos"

        df_clasificado["Score_Oportunidad"] = df_clasificado.apply(calcular_score_oportunidad, axis=1)
        df_geocoded = geocode_addresses(df_clasificado)
        
        df_geocoded["Distancia_Desde_Base"] = df_geocoded.apply(
            lambda row: calcular_distancia(BASE_LAT, BASE_LNG, row["Latitud"], row["Longitud"]) 
            if pd.notna(row["Latitud"]) and pd.notna(row["Longitud"]) else None, axis=1
        )

        df_geocoded["Frecuencia_Visita_Dias"] = df_geocoded.apply(calcular_frecuencia_visita, axis=1)
        df_geocoded["Score_Prioridad_Final"] = df_geocoded.apply(calcular_prioridad_final, axis=1)
        df_clustered = cluster_clients(df_geocoded, num_clusters)
        
        output_path = os.path.join(temp_dir, "resultados.xlsx")
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # 1. Hoja My Maps
            create_my_maps_sheet(df_clustered, writer)

            # 2. Hoja de Detalle de Clientes
            df_detalle_clientes = df_clustered[[
                "Razón Social", "Categoria_ABC", "Score_Oportunidad", 
                "Score_Prioridad_Final", "Cluster", "Venta_Total_Anual", 
                "Frecuencia_Visita_Dias", "Distancia_Desde_Base", "Latitud", "Longitud"
            ]].copy()
            df_detalle_clientes.rename(columns={
                "Razón Social": "Cliente", "Categoria_ABC": "Categoría ABC",
                "Score_Oportunidad": "Score Oportunidad", "Score_Prioridad_Final": "Score Prioridad",
                "Venta_Total_Anual": "Venta Total Anual", "Frecuencia_Visita_Dias": "Frecuencia Visita (días)",
                "Distancia_Desde_Base": "Distancia Base (km)"
            }, inplace=True)
            df_detalle_clientes.to_excel(writer, sheet_name="Detalle de Clientes", index=False)
            apply_styles_and_autofit(writer, "Detalle de Clientes", df_detalle_clientes)

            # 3. Hoja de Dashboard
            create_dashboard_sheet(df_clustered, writer)

            # 4. Hoja de Datos Unificados
            df_unificado_formateado = df_clustered.rename(columns={"Razón Social": "Cliente"})
            df_unificado_formateado.to_excel(writer, sheet_name="Datos Unificados", index=False)
            apply_styles_and_autofit(writer, "Datos Unificados", df_unificado_formateado)

            # 5. Hojas por Cluster
            for cluster_id in sorted(df_clustered["Cluster"].unique()):
                sheet_name = "Sin Cluster" if cluster_id == -1 else f"Cluster {int(cluster_id) + 1}"
                
                df_cluster = df_clustered[df_clustered["Cluster"] == cluster_id].sort_values(by="Score_Prioridad_Final", ascending=False)
                df_cluster_display = df_cluster[[
                    "Razón Social", "Categoria_ABC", "Score_Oportunidad", 
                    "Score_Prioridad_Final", "Venta_Total_Anual", 
                    "Frecuencia_Visita_Dias", "Distancia_Desde_Base"
                ]].copy()
                df_cluster_display.rename(columns={
                    "Razón Social": "Cliente", "Categoria_ABC": "Categoría ABC",
                    "Score_Oportunidad": "Score Oportunidad", "Score_Prioridad_Final": "Score Prioridad",
                    "Venta_Total_Anual": "Venta Total Anual", "Frecuencia_Visita_Dias": "Frecuencia Visita (días)",
                    "Distancia_Desde_Base": "Distancia Base (km)"
                }, inplace=True)
                df_cluster_display.to_excel(writer, sheet_name=sheet_name, index=False)
                apply_styles_and_autofit(writer, sheet_name, df_cluster_display)

        return send_file(output_path, as_attachment=True, download_name="resultados_procesados.xlsx")
        
    except Exception as e:
        print(f"Error en procesar_datos: {str(e)}")
        print(traceback.format_exc())
        return jsonify({"error": f"Error interno del servidor: {str(e)}"}), 500

